"""
Tests for the general functioning of the library
"""
import os
import sys
import subprocess
import shutil

import pytest
import numpy as np
from pysd import read_vensim

import excels2vensim as e2v


encoding_stdout = sys.stdout.encoding or "utf-8"
encoding_stderr = sys.stderr.encoding or "utf-8"

def test_constants(tmp_path, _root):

    os.chdir(_root / "tmp_dir")
    # copy original file without data
    shutil.copy2(_root / 'original_files' / 'inputs.xlsx',
                 'inputs.xlsx')

    obj0 = e2v.Constants(
        'var0',
        [],
        'A24',
        'This is my 0 dim variable',
        'Twh',
        'inputs.xlsx',
        'Region1')

    obj0.get_vensim()

    e2v.Subscripts.set({'source': ['Gas', 'Oil', 'Coal']})

    obj1 = e2v.Constants(
        'q_row',
        ['source'],
        'A24',
        'This is my variable q_row',
        'Twh',
        'inputs.xlsx',
        'Region1')

    obj1.add_dimension('source', 'col')

    obj1.get_vensim()

    obj2 = e2v.Constants(
        'q_col',
        ['source'],
        'B18',
        'This is my variable q_col',
        'Twh',
        'inputs.xlsx',
        'Region1')

    obj2.add_dimension('source', 'row')

    obj2.get_vensim()

    # add more subscripts (check update method)
    e2v.Subscripts.update({
        'sector': ['A', 'B', 'C', 'D'],
        'region': ['Region1', 'Region2', 'Region3', 'Region4'],
        'out': ['Elec', 'Heat', 'Solid', 'Liquid']})

    result = e2v.load_from_json(_root / 'jsons' / 'constants.json')

    with open(_root / 'original_files' / 'model_constants.mdl') as file:
        model = file.read()

    model += result

    with open('model_constants.mdl', 'w') as file:
        file.write(model)

    model = read_vensim('model_constants.mdl')

    var = model._external_elements[0]()

    assert var.dims == ('source', 'sector', 'region', 'out')
    assert not np.any(np.isnan(var.values))


def test_data(tmp_path, _root):

    os.chdir(_root / "tmp_dir")
    # copy original file without cellranges
    shutil.copy2(_root / 'original_files' / 'inputs_data.xlsx',
                 'inputs_data.xlsx')

    # read subscripts from a mdl file
    e2v.Subscripts.read(_root / 'subscripts' / 'data.mdl')

    # test load_from_json
    result = e2v.load_from_json(_root / 'jsons' / 'data.json')

    with open(_root / 'original_files' / 'model_data.mdl') as file:
        model = file.read()

    model += result

    with open('model_data.mdl', 'w') as file:
        file.write(model)

    model = read_vensim('model_data.mdl')

    var = model._external_elements[0].data

    assert var.dims == ('time', 'AGE COHORT', 'REGIONS', 'GENDER')
    assert not np.any(np.isnan(var.values))

    # assert difference between specific subscripts
    assert np.all(var.sel(time=2005).values
                  != var.sel(time=2015).values)
    assert np.all(var.sel(REGIONS='EU27').values
                  != var.sel(REGIONS='CHINA').values)
    assert np.all(var.sel(GENDER='female').values
                  != var.sel(GENDER='male').values)
    assert np.all(var.loc[{'AGE COHORT': '"50-54"'}].values
                  != var.loc[{'AGE COHORT': '"20-24"'}].values)

    # assert some numeric values
    assert int(var.loc['2005', '"0-4"', "EU27", "female"]) == 10683786
    assert int(var.loc['2007', '"10-14"', "LATAM", "male"]) == 15310480


def test_lookup(tmp_path, _root):

    os.chdir(_root / "tmp_dir")
    # copy original file without cellranges
    shutil.copy2(_root / 'original_files' / 'inputs_data2.xlsx',
                 _root / 'tmp_dir' / 'inputs_data2.xlsx')

    out_dir = "lookup_output.txt"

    subs_dir = str(_root / "subscripts" / "data_subscripts.json")

    conf_dir = str(_root / "jsons" / "lookups.json")

    # test command line
    subprocess.run([
        "python3", "-m", "excels2vensim",
        "--output-file=" + out_dir,
        subs_dir, conf_dir])

    with open(out_dir) as file:
        result = file.read()

    with open(_root / 'original_files' / 'model_data.mdl') as file:
        model = file.read()

    model += result

    with open('model_lookup.mdl', 'w') as file:
        file.write(model)

    model = read_vensim('model_lookup.mdl')

    var = model._external_elements[0].data

    assert var.dims == ('lookup_dim', 'GENDER', 'AGE COHORT', 'REGIONS')
    assert not np.any(np.isnan(var.values))

    # assert difference between specific subscripts
    assert np.all(var.sel(lookup_dim=2005).values
                  != var.sel(lookup_dim=2015).values)
    assert np.all(var.sel(REGIONS='EU27').values
                  != var.sel(REGIONS='CHINA').values)
    assert np.all(var.sel(GENDER='female').values
                  != var.sel(GENDER='male').values)
    assert np.all(var.loc[{'AGE COHORT': '"50-54"'}].values
                  != var.loc[{'AGE COHORT': '"20-24"'}].values)

    # assert some numeric values
    assert int(var.loc['2005', "female", '"0-4"', "EU27", ]) == 10683786
    assert int(var.loc['2007', "male", '"10-14"', "LATAM", ]) == 15310480


def test_non_valid_chars(tmp_path, _root):

    os.chdir(_root / "tmp_dir")
    # copy original file without data
    shutil.copy2(_root / 'original_files' / 'inputs.xlsx',
                 'inputs_nvc.xlsx')

    expected = r"The name of the variable 'my q row\$' has special characters"\
               + r"\. 'my_q_row' will be used for cellrange names."

    e2v.Subscripts.set({
        'source': ['Gas', 'Oil', 'Coal'],
        'sector': ['A', 'B', 'C', 'D'],
        'region': ['Region1', 'Region2', 'Region3', 'Region4'],
        'out': ['"  Elec/el"', 'Heat', ' "Solid$"', 'Liquid']})

    # invalid var name
    with pytest.warns(UserWarning, match=expected):
        obj1 = e2v.Constants(
            'my q row$',
            ['source'],
            'A24',
            'This is my variable my q row$',
            'Twh',
            'inputs_nvc.xlsx',
            'Region1')

    obj1.add_dimension('source', 'col')

    out = obj1.get_vensim()

    assert "my q row$[source]=" in out
    assert "CONSTANTS('inputs_nvc.xlsx', 'Region1', 'my_q_row')" in out

    # invalid dim name
    with pytest.warns(UserWarning) as record:
        out = e2v.load_from_json(_root / 'jsons' / 'constants_nvc.json')

    record = [str(r.message) for r in record]

    expected = [
        "The name of the subscript '\"  Elec/el\"' has special characters. "
        + "'Elec_el' will be used for cellrange names.",
        "The name of the subscript '\"Solid$\"' has special characters. "
        + "'Solid' will be used for cellrange names."]

    for message in expected:
        assert message in record

    assert "share_energy[source, sector, Region3, \"  Elec/el\"]=\n\t"\
        + "GET_DIRECT_CONSTANTS('inputs_nvc.xlsx', 'Region3',"\
        + " 'share_energy_Elec_el') ~~|"\
        in out

    assert "share_energy[source, sector, Region1, \"Solid$\"]=\n\t"\
        + "GET_DIRECT_CONSTANTS('inputs_nvc.xlsx', 'Region1',"\
        + " 'share_energy_Solid') ~~|"\
        in out

    # invalid series name
    # copy original file without cellranges
    shutil.copy2(_root / 'original_files' / 'inputs_data.xlsx',
                 'inputs_data_nvs.xlsx')

    # read subscripts from a mdl file
    e2v.Subscripts.read(_root / 'subscripts' / 'data.mdl')

    expected = r"The name of the interpolation dimension 'my time\$'"\
               + r" has special characters\. 'my_time' will be used for "\
               + r"cellrange names\."
    with pytest.warns(UserWarning, match=expected):
        out = e2v.load_from_json(_root / 'jsons' / 'data_nvseries.json')

    assert "DATA('inputs_data_nvs.xlsx', 'GPH', 'my_time',"\
        in out


def test_data_with_keywords(tmp_path, _root):
    """
    Test for DATA with 'HOLD BACKWARD' keyword
    """

    os.chdir(_root / "tmp_dir")
    # copy original file without cellranges
    shutil.copy2(_root / 'original_files' / 'inputs_data.xlsx',
                 'inputs_data_k.xlsx')

    # read subscripts from a mdl file
    e2v.Subscripts.read(_root / 'subscripts' / 'data.mdl')

    # test load_from_json
    result = e2v.load_from_json(_root / 'jsons' / 'data_keywords.json')

    assert "]:=" not in result
    assert "]:HOLD BACKWARD::=" in result

    # non valid keyword
    expected = "\ninterp must be 'interpolate', 'raw', "\
               + "'hold backward' or 'look forward'"

    # invalid var name
    with pytest.raises(ValueError, match=expected):
        e2v.Data(
            'pop',
            ['REGION'],
            'A24',
            '',
            '',
            '',
            '',
            interp='non_valid')


@pytest.mark.skipif(
    sys.platform.startswith("win"),
    reason="not working on Windows, stdout produces extra line-breaks in"
           " GHActions. Locally is working fine."
)
def test_dimensionless_data(tmp_path, _root):

    os.chdir(_root / "tmp_dir")
    # copy original file without cellranges
    shutil.copy2(_root / 'original_files' / 'inputs_data2.xlsx',
                 _root / 'tmp_dir' / 'inputs_dmnl.xlsx')

    subs_dir = str(_root / "subscripts" / "data_subscripts.json")

    conf_dir = str(_root / "jsons" / "dimensionless.json")

    # test command line
    out = subprocess.run([
        "python3", "-m", "excels2vensim",
        subs_dir, conf_dir], capture_output=True)

    stdout = out.stdout.decode(encoding_stdout)

    assert "population_lookup=\n\t"\
        "GET_DIRECT_LOOKUPS('../tmp_dir/inputs_dmnl.xlsx', 'EU27',"\
        " 'time', 'population_lookup')" in stdout

    assert "population_data:=\n\t"\
        "GET_DIRECT_DATA('../tmp_dir/inputs_dmnl.xlsx', 'EU27',"\
        " 'time', 'population_data')" in stdout

    assert "population_constant=\n\t"\
        "GET_DIRECT_CONSTANTS('../tmp_dir/inputs_dmnl.xlsx', 'EU27',"\
        " 'population_constant')" in stdout


def test_force(tmp_path, _root):
    """
    Test for force option
    """
    from openpyxl import load_workbook

    os.chdir(_root / "tmp_dir")
    # copy original file without data
    shutil.copy2(_root / "original_files" / "inputs.xlsx",
                 "inputs_force.xlsx")

    element_dict = {
        "my_var": {
            "type": "constants",
            "force": False,
            "loading": "DIRECT",
            "dims": [],
            "cell": "A10",
            "description": "",
            "units": "",
            "file": "inputs_force.xlsx",
            "sheet": "Region1",
            "dimensions": {}
        }}

    e2v.execute(element_dict)

    wb = load_workbook("inputs_force.xlsx")
    ws = wb["Region1"]
    assert "my_var" in ws.defined_names
    assert ws.defined_names.get("my_var").attr_text == 'Region1!$A$10:$A$10'
    wb.close()

    element_dict["my_var"]["cell"] = "B34"

    expected = r"Trying to write a cellrange with name 'my_var' at "\
               r"'Region1!\$B\$34:\$B\$34'. However, 'my_var' already exist "\
               r"in 'Region1!\$A\$10:\$A\$10'\nUse force=True to overwrite it."
    with pytest.raises(ValueError, match=expected):
        e2v.execute(element_dict)

    element_dict["my_var"]["force"] = True

    e2v.execute(element_dict)

    wb = load_workbook("inputs_force.xlsx")
    ws = wb["Region1"]
    assert "my_var" in ws.defined_names
    assert ws.defined_names.get("my_var").attr_text == 'Region1!$B$34:$B$34'
    wb.close()
